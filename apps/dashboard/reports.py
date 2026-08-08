"""
apps/dashboard/reports.py
Generacion de reportes Excel para contabilidad (RF-28), con openpyxl.
Ventas, liquidaciones y horas extras, filtrables por periodo.
"""

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color='3E2B1F', end_color='3E2B1F', fill_type='solid')
HEADER_FONT = Font(color='FAF7F2', bold=True)
TOTAL_FONT = Font(bold=True)


def _new_workbook(sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws


def _write_header(ws, headers, row=1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'


def _autosize_columns(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 42)


def _http_response(wb, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def build_sales_report(start_date, end_date):
    from datetime import datetime, time

    from django.utils import timezone

    from apps.store.models import Order

    wb, ws = _new_workbook('Ventas')
    headers = ['N° Pedido', 'Fecha', 'Cliente', 'Email', 'RUT', 'Estado', 'Neto', 'IVA', 'Total']
    _write_header(ws, headers)

    # Se filtra por rango de datetimes en zona local, no por created_at__date:
    # con USE_TZ=True los datetimes se guardan en UTC, y en Chile (UTC-4) un
    # pedido de las 21:00 caeria en el dia siguiente segun UTC.
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.combine(start_date, time.min), tz)
    end_dt = timezone.make_aware(datetime.combine(end_date, time.max), tz)

    orders = Order.objects.filter(
        created_at__gte=start_dt, created_at__lte=end_dt,
    ).order_by('created_at')

    row = 2
    for order in orders:
        ws.cell(row=row, column=1, value=order.order_number)
        ws.cell(row=row, column=2, value=timezone.localtime(order.created_at).strftime('%d-%m-%Y %H:%M'))
        ws.cell(row=row, column=3, value=order.contact_name)
        ws.cell(row=row, column=4, value=order.contact_email)
        ws.cell(row=row, column=5, value=order.contact_rut)
        ws.cell(row=row, column=6, value=order.get_status_display())
        ws.cell(row=row, column=7, value=float(order.net_total))
        ws.cell(row=row, column=8, value=float(order.iva_total))
        ws.cell(row=row, column=9, value=float(order.total))
        row += 1

    if row > 2:
        ws.cell(row=row, column=6, value='TOTAL').font = TOTAL_FONT
        for col in ('G', 'H', 'I'):
            cell = ws.cell(row=row, column=ord(col) - ord('A') + 1, value=f'=SUM({col}2:{col}{row - 1})')
            cell.font = TOTAL_FONT

    _autosize_columns(ws)
    filename = f'ventas_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx'
    return _http_response(wb, filename)


def build_payroll_report(start_date, end_date):
    from apps.hr.models import PayrollRecord

    wb, ws = _new_workbook('Liquidaciones')
    headers = [
        'Trabajador', 'RUT', 'Periodo', 'Sueldo base', 'Horas extras', 'Bonos',
        'Imponible', 'AFP', '% AFP', 'Cotización AFP', 'Salud (7%)',
        'Anticipos', 'Descuentos', 'Líquido', 'Estado',
    ]
    _write_header(ws, headers)

    records = PayrollRecord.objects.select_related('worker__user').filter(
        period__gte=start_date.replace(day=1), period__lte=end_date,
    ).order_by('period', 'worker__user__first_name')

    row = 2
    for r in records:
        values = [
            str(r.worker), r.worker.rut, r.period.strftime('%m-%Y'),
            float(r.base_salary), float(r.overtime_total), float(r.bonuses_total),
            float(r.taxable_income), r.afp_name, float(r.afp_percentage),
            float(r.afp_amount), float(r.health_amount), float(r.advances_total),
            float(r.deductions_total), float(r.net_pay), r.get_status_display(),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1

    if row > 2:
        ws.cell(row=row, column=8, value='TOTAL').font = TOTAL_FONT
        for col in ('J', 'K', 'N'):
            cell = ws.cell(row=row, column=ord(col) - ord('A') + 1, value=f'=SUM({col}2:{col}{row - 1})')
            cell.font = TOTAL_FONT

    _autosize_columns(ws)
    filename = f'liquidaciones_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx'
    return _http_response(wb, filename)


def build_overtime_report(start_date, end_date):
    from apps.hr.models import OvertimeRecord

    wb, ws = _new_workbook('Horas extras')
    headers = ['Trabajador', 'RUT', 'Fecha', 'Horas', 'Valor hora ordinaria', 'Valor hora extra', 'Total', 'Estado']
    _write_header(ws, headers)

    records = OvertimeRecord.objects.select_related('worker__user').filter(
        date__gte=start_date, date__lte=end_date,
    ).order_by('date')

    row = 2
    for r in records:
        values = [
            str(r.worker), r.worker.rut, r.date.strftime('%d-%m-%Y'),
            float(r.hours), float(r.hourly_rate), float(r.overtime_rate),
            float(r.total), r.get_status_display(),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=value)
        row += 1

    if row > 2:
        ws.cell(row=row, column=6, value='TOTAL').font = TOTAL_FONT
        cell = ws.cell(row=row, column=7, value=f'=SUM(G2:G{row - 1})')
        cell.font = TOTAL_FONT

    _autosize_columns(ws)
    filename = f'horas_extras_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx'
    return _http_response(wb, filename)
